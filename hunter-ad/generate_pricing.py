"""
Hunter Pricing Plans — Excel Generator
Run: python generate_pricing.py
Output: hunter_pricing_plans.xlsx
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ────────────────────────────────────────────────────────────────────
BLACK   = "09090B"
RED     = "DC2626"
WHITE   = "FFFFFF"
Z900    = "18181B"
Z800    = "27272A"
Z700    = "3F3F46"
Z400    = "A1A1AA"
Z200    = "E4E4E7"
GREEN   = "22C55E"
AMBER   = "F59E0B"

def cell_style(ws, row, col,
               value="", bold=False, size=11, color=WHITE,
               bg=None, align="left", wrap=False, border=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Inter", bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if border:
        thin = Side(style="thin", color=Z700)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c

# ── Sheet 1: Pricing Overview ─────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Pricing Overview"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 36
for col in ["B","C","D","E","F"]:
    ws1.column_dimensions[col].width = 22

# Row heights
ws1.row_dimensions[1].height = 60
ws1.row_dimensions[2].height = 28
ws1.row_dimensions[3].height = 36
ws1.row_dimensions[4].height = 24
ws1.row_dimensions[5].height = 24
for r in range(6, 60):
    ws1.row_dimensions[r].height = 22

# Header banner
ws1.merge_cells("A1:F1")
h = ws1["A1"]
h.value = "HUNTER  —  Pricing Plans  |  B2B Lead Intelligence · Kenya"
h.font = Font(name="Inter", bold=True, size=18, color=WHITE)
h.fill = PatternFill("solid", fgColor=BLACK)
h.alignment = Alignment(horizontal="center", vertical="center")

ws1.merge_cells("A2:F2")
sub = ws1["A2"]
sub.value = "Effective 2026  ·  All plans include 7-day free trial  ·  No credit card required for trial"
sub.font = Font(name="Inter", size=11, color=Z400)
sub.fill = PatternFill("solid", fgColor=BLACK)
sub.alignment = Alignment(horizontal="center", vertical="center")

# Plan headers
plans = [
    ("",        BLACK, WHITE, ""),
    ("TRIAL",   Z900,  Z400,  "7 days"),
    ("SOLO",    Z900,  WHITE, "KES 2,500/mo"),
    ("GROWTH",  RED,   WHITE, "KES 6,500/mo"),
    ("AGENCY",  Z900,  WHITE, "KES 15,000/mo"),
    ("ENTERPRISE", BLACK, Z400, "Custom"),
]

ws1.row_dimensions[3].height = 42
ws1.row_dimensions[4].height = 26

for ci, (name, bg, fg, price) in enumerate(plans, 1):
    cell_style(ws1, 3, ci, name, bold=True, size=14, color=fg, bg=bg, align="center")
    cell_style(ws1, 4, ci, price, bold=False, size=12, color=AMBER if name=="GROWTH" else Z400, bg=bg, align="center")

# USD equivalents row
usd_row = ["USD (approx.)", "Free", "~$19/mo", "~$49/mo", "~$115/mo", "Custom"]
ws1.row_dimensions[5].height = 20
for ci, v in enumerate(usd_row, 1):
    cell_style(ws1, 5, ci, v, size=10, color=Z400, bg=Z900 if ci>1 else BLACK, align="center")

# Annual pricing row
ann_row = ["Annual (2 months free)", "—", "KES 25,000/yr", "KES 65,000/yr", "KES 150,000/yr", "Custom"]
ws1.row_dimensions[6].height = 20
for ci, v in enumerate(ann_row, 1):
    cell_style(ws1, 6, ci, v, size=10, color=GREEN if "KES" in v else Z400, bg=Z900 if ci>1 else BLACK, align="center")

# Section + feature rows
features = [
    # (section_header, values per plan: trial/solo/growth/agency/enterprise)
    ("── USAGE ──", None),
    ("Monthly leads",          ["200",    "200",    "1,000",   "Unlimited", "Unlimited"]),
    ("Team seats",             ["1",      "1",      "1",       "5",         "Unlimited"]),
    ("Bulk enrichment",        ["✓",      "✓",      "✓",       "✓",         "✓"]),

    ("── ENRICHMENT ──", None),
    ("Website crawl (email, WA, socials)", ["✓","✓","✓","✓","✓"]),
    ("Booking & live chat detection",      ["✓","✓","✓","✓","✓"]),
    ("About page contact discovery",       ["—","—","✓","✓","✓"]),
    ("Instagram bio contact extraction",   ["—","—","✓","✓","✓"]),
    ("Custom signal detection",            ["✓","✓","✓","✓","✓"]),

    ("── ENRICHMENT MODES ──", None),
    ("Available modes",        ["1 (General)","1 (General)","All 8","All 8","All 8 + Custom"]),
    ("Mode: Digital / Agency", ["—","—","✓","✓","✓"]),
    ("Mode: Insurance Sales",  ["—","—","✓","✓","✓"]),
    ("Mode: Fintech / Lending",["—","—","✓","✓","✓"]),
    ("Mode: Solar / Energy",   ["—","—","✓","✓","✓"]),
    ("Mode: Telecom",          ["—","—","✓","✓","✓"]),
    ("Mode: Healthcare",       ["—","—","✓","✓","✓"]),
    ("Mode: Recruiter",        ["—","—","✓","✓","✓"]),

    ("── AI FEATURES ──", None),
    ("Gemini lead scoring (0-100)",    ["✓","✓","✓","✓","✓"]),
    ("AI scoring reasoning",           ["✓","✓","✓","✓","✓"]),
    ("Mode-aware scoring frame",       ["—","—","✓","✓","✓"]),
    ("Historical feedback calibration",["—","—","✓","✓","✓"]),
    ("WhatsApp opener generation",     ["✓","✓","✓","✓","✓"]),
    ("Email opener generation",        ["✓","✓","✓","✓","✓"]),
    ("Mode-aware outreach angle",      ["—","—","✓","✓","✓"]),

    ("── DISCOVERY ──", None),
    ("Google Maps scraping",           ["✓","✓","✓","✓","✓"]),
    ("OpenStreetMap scraping",         ["✓","✓","✓","✓","✓"]),
    ("26 supported verticals",         ["✓","✓","✓","✓","✓"]),
    ("10 supported cities",            ["✓","✓","✓","✓","✓"]),
    ("Bulk score all leads",           ["✓","✓","✓","✓","✓"]),

    ("── PIPELINE & FEEDBACK ──", None),
    ("Pipeline / kanban view",         ["—","—","✓","✓","✓"]),
    ("Lead outcome feedback",          ["✓","✓","✓","✓","✓"]),
    ("Feedback loop scoring",          ["—","—","✓","✓","✓"]),
    ("Contact accuracy tracking",      ["—","—","✓","✓","✓"]),
    ("Follow-up date scheduling",      ["—","—","✓","✓","✓"]),
    ("Lead notes",                     ["✓","✓","✓","✓","✓"]),

    ("── EXPORT & INTEGRATION ──", None),
    ("CSV export",                     ["—","—","✓","✓","✓"]),
    ("REST API access",                ["—","—","—","Limited","Full"]),
    ("Webhook support",                ["—","—","—","✓","✓"]),
    ("CRM integration (Zapier)",       ["—","—","—","✓","✓"]),
    ("White-label exports",            ["—","—","—","✓","✓"]),

    ("── CORPORATE / TEAM ──", None),
    ("Corporate account type",         ["—","—","—","✓","✓"]),
    ("Team workspace",                 ["—","—","—","✓","✓"]),
    ("Shared lead pool",               ["—","—","—","✓","✓"]),
    ("Team analytics",                 ["—","—","—","✓","✓"]),
    ("Role-based access (admin/member)",["—","—","—","✓","✓"]),

    ("── SUPPORT ──", None),
    ("Email support",                  ["—","✓","✓","✓","✓"]),
    ("WhatsApp support",               ["—","—","✓","✓","✓"]),
    ("Priority support SLA (24h)",     ["—","—","—","✓","✓"]),
    ("Dedicated account manager",      ["—","—","—","✓","✓"]),
    ("Custom onboarding session",      ["—","—","—","—","✓"]),
    ("SLA uptime guarantee (99.9%)",   ["—","—","—","—","✓"]),
]

row = 7
for item in features:
    label, vals = item
    if vals is None:
        # Section header
        ws1.row_dimensions[row].height = 28
        ws1.merge_cells(f"A{row}:F{row}")
        cell_style(ws1, row, 1, label, bold=True, size=10, color=RED, bg=Z900, align="left")
        row += 1
        continue

    ws1.row_dimensions[row].height = 22
    cell_style(ws1, row, 1, label, size=10, color=Z200, bg=Z900 if row%2==0 else BLACK, align="left")
    plan_bgs = [Z900, Z900, RED, Z900, BLACK]
    for ci, (v, bg) in enumerate(zip(vals, plan_bgs), 2):
        color = GREEN if v == "✓" else (Z400 if v == "—" else WHITE)
        cell_style(ws1, row, ci, v, bold=(v=="✓"), size=10,
                   color=color, bg=bg if ci==4 else (Z900 if row%2==0 else BLACK),
                   align="center")
    row += 1

# ── Sheet 2: Detailed Pricing Table ──────────────────────────────────────────
ws2 = wb.create_sheet("Plan Details & Revenue")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 18
ws2.column_dimensions["E"].width = 18
ws2.column_dimensions["F"].width = 20
ws2.column_dimensions["G"].width = 24

ws2.merge_cells("A1:G1")
ws2["A1"].value = "HUNTER — Plan Revenue Model"
ws2["A1"].font = Font(name="Inter", bold=True, size=16, color=WHITE)
ws2["A1"].fill = PatternFill("solid", fgColor=BLACK)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 48

headers = ["Plan", "Monthly (KES)", "Monthly (USD)", "Annual (KES)", "Annual (USD)", "Trial Period", "Notes"]
for ci, h in enumerate(headers, 1):
    cell_style(ws2, 2, ci, h, bold=True, size=11, color=WHITE, bg=Z800, align="center")
ws2.row_dimensions[2].height = 28

plan_rows = [
    ["Trial",      "0",          "$0",     "N/A",      "N/A",    "7 days",    "Full Starter features. Converts to Solo."],
    ["Solo",       "2,500",      "~$19",   "25,000",   "~$192",  "7 days",    "Ideal for solo sales reps & freelancers"],
    ["Growth",     "6,500",      "~$49",   "65,000",   "~$500",  "7 days",    "Best for active field sales teams"],
    ["Agency",     "15,000",     "~$115",  "150,000",  "~$1,154","7 days",    "5 seats. White-label. API access."],
    ["Enterprise", "Custom",     "Custom", "Custom",   "Custom", "Negotiable","Unlimited. SLA. Custom integrations."],
]

for ri, row_data in enumerate(plan_rows, 3):
    bg_map = ["", Z900, Z900, RED, Z900, BLACK]
    for ci, val in enumerate(row_data, 1):
        bg = RED if ri == 5 and ci > 1 else (Z900 if ri%2==0 else BLACK)
        vc = AMBER if "KES" in val or "$" in val else (GREEN if val == "7 days" else WHITE)
        cell_style(ws2, ri, ci, val, size=11, color=vc if ci > 1 else WHITE, bg=bg, align="center" if ci > 1 else "left")
    ws2.row_dimensions[ri].height = 26

# Revenue projections
ws2.row_dimensions[9].height = 32
ws2.merge_cells("A9:G9")
cell_style(ws2, 9, 1, "── REVENUE PROJECTIONS (Monthly Recurring Revenue) ──", bold=True, size=11, color=RED, bg=Z900)

proj_headers = ["Scenario", "Solo Users", "Growth Users", "Agency Users", "Total MRR (KES)", "Total MRR (USD)", "Notes"]
for ci, h in enumerate(proj_headers, 1):
    cell_style(ws2, 10, ci, h, bold=True, size=10, color=WHITE, bg=Z800, align="center")
ws2.row_dimensions[10].height = 24

scenarios = [
    ["Month 3 (Launch)",    "20",   "5",   "1",   "=B11*2500+C11*6500+D11*15000", "=E11/130", "Early adopter traction"],
    ["Month 6",             "60",   "20",  "5",   "=B12*2500+C12*6500+D12*15000", "=E12/130", "Growth phase"],
    ["Month 12 (Year 1)",   "150",  "60",  "15",  "=B13*2500+C13*6500+D13*15000", "=E13/130", "~KES 1.1M/mo target"],
    ["Month 18 (Scale)",    "400",  "150", "40",  "=B14*2500+C14*6500+D14*15000", "=E14/130", "Agency + corporate growth"],
    ["Month 24 (Target)",   "1000", "400", "100", "=B15*2500+C15*6500+D15*15000", "=E15/130", "KES 5.7M/mo = ~$44k USD"],
]

for ri, row_data in enumerate(scenarios, 11):
    for ci, val in enumerate(row_data, 1):
        is_formula = val.startswith("=")
        bg = Z900 if ri%2==0 else BLACK
        color = AMBER if ci in [5,6] else (WHITE if ci == 1 else Z200)
        c = ws2.cell(row=ri, column=ci, value=val)
        c.font = Font(name="Inter", size=10, bold=(ci in [5,6]), color=color)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center" if ci>1 else "left", vertical="center")
        if is_formula and ci == 5:
            c.number_format = '#,##0'
        if is_formula and ci == 6:
            c.number_format = '$#,##0'
    ws2.row_dimensions[ri].height = 22

# ── Sheet 3: Trial & Conversion ───────────────────────────────────────────────
ws3 = wb.create_sheet("Trial & Conversion")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 32
ws3.column_dimensions["B"].width = 24
ws3.column_dimensions["C"].width = 24
ws3.column_dimensions["D"].width = 30

ws3.merge_cells("A1:D1")
ws3["A1"].value = "HUNTER — Trial Window & Conversion Strategy"
ws3["A1"].font = Font(name="Inter", bold=True, size=16, color=WHITE)
ws3["A1"].fill = PatternFill("solid", fgColor=BLACK)
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 48

trial_data = [
    ("Trial Duration",          "7 days",                "From account creation",   "No credit card required"),
    ("Trial Plan Level",        "Starter (200 leads)",   "Full Solo feature set",   "Contact discovery excluded"),
    ("Trial Reminder (Day 5)",  "Email: 2 days left",    "In-app banner shown",     "Soft upgrade prompt"),
    ("Trial Reminder (Day 7)",  "Email: expires today",  "Hard paywall at midnight","Downgrade to 10 leads/mo"),
    ("Post-trial (unpaid)",     "Read-only mode",        "Can view existing leads", "Cannot scrape or enrich"),
    ("Grace period",            "3 days",                "Post-trial read-only",    "One-click reactivation"),
    ("Trial to Paid target",    "25% conversion",        "Industry avg: 15-20%",   "Onboarding call helps"),
    ("Corp trial extension",    "+7 days (14 total)",    "For verified companies",  "Sales-assisted onboarding"),
]

headers3 = ["Attribute", "Value", "Detail", "Notes"]
for ci, h in enumerate(headers3, 1):
    cell_style(ws3, 2, ci, h, bold=True, size=11, color=WHITE, bg=Z800, align="center")
ws3.row_dimensions[2].height = 28

for ri, row_data in enumerate(trial_data, 3):
    for ci, val in enumerate(row_data, 1):
        bg = Z900 if ri%2==0 else BLACK
        color = GREEN if ri == 10 else (AMBER if "Day 7" in val or "expires" in val else (WHITE if ci == 1 else Z200))
        cell_style(ws3, ri, ci, val, size=10, color=WHITE if ci==1 else Z200, bg=bg, align="left")
    ws3.row_dimensions[ri].height = 22

# Email sequence schedule
ws3.row_dimensions[12].height = 32
ws3.merge_cells("A12:D12")
cell_style(ws3, 12, 1, "── TRIAL EMAIL SEQUENCE ──", bold=True, size=11, color=RED, bg=Z900)

email_seq = [
    ("Day 0 — Welcome",    "Subject: Your Hunter trial is live 🎯",   "CTA: Run your first scrape → /discover", "Tone: Excited, helpful"),
    ("Day 2 — Tip",        "Subject: Found your first leads?",          "CTA: Try AI scoring → /leads",          "Feature: score + opener"),
    ("Day 5 — Urgency",    "Subject: 2 days left on your trial",       "CTA: Upgrade now — lock in your rate",  "Show: Solo plan benefits"),
    ("Day 7 — Last call",  "Subject: Your trial ends tonight",         "CTA: Upgrade to keep your leads",       "Urgency + loss aversion"),
    ("Day 8 — Re-engage",  "Subject: Your leads are still here",       "CTA: Come back — upgrade to access",    "For non-converters"),
    ("Day 14 — Win-back",  "Subject: Hunter has new features for you", "CTA: Start a new trial / upgrade",      "30% off first month offer"),
]

headers_email = ["Timing", "Subject Line", "Primary CTA", "Notes"]
for ci, h in enumerate(headers_email, 1):
    cell_style(ws3, 13, ci, h, bold=True, size=10, color=WHITE, bg=Z800, align="center")
ws3.row_dimensions[13].height = 24

for ri, row_data in enumerate(email_seq, 14):
    for ci, val in enumerate(row_data, 1):
        bg = Z900 if ri%2==0 else BLACK
        color = RED if "Last call" in val or "Urgency" in val else (WHITE if ci==1 else Z200)
        cell_style(ws3, ri, ci, val, size=10, color=color if ci==1 else Z200, bg=bg, align="left")
    ws3.row_dimensions[ri].height = 22

# ── Sheet 4: Corporate Plans ──────────────────────────────────────────────────
ws4 = wb.create_sheet("Corporate Accounts")
ws4.sheet_view.showGridLines = False
for col in ["A","B","C","D","E"]:
    ws4.column_dimensions[col].width = 28

ws4.merge_cells("A1:E1")
ws4["A1"].value = "HUNTER — Corporate Account Structure"
ws4["A1"].font = Font(name="Inter", bold=True, size=16, color=WHITE)
ws4["A1"].fill = PatternFill("solid", fgColor=BLACK)
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 48

corp_features = [
    ("Feature",                      "Individual",  "Agency (5 seats)", "Enterprise (10+)", "Enterprise+"),
    ("Account type",                 "Personal",    "Corporate",        "Corporate",        "Corporate"),
    ("Registered company required",  "No",          "Optional",         "Yes",              "Yes"),
    ("Max seats",                    "1",           "5",                "10",               "Unlimited"),
    ("Lead pool sharing",            "No",          "Yes",              "Yes",              "Yes"),
    ("Shared outreach history",      "No",          "Yes",              "Yes",              "Yes"),
    ("Admin + member roles",         "No",          "Yes",              "Yes",              "Yes"),
    ("Billing contact separate",     "No",          "Yes",              "Yes",              "Yes"),
    ("Invoicing (NET-30)",           "No",          "No",               "Yes",              "Yes"),
    ("M-PESA Business payment",      "Yes",         "Yes",              "Yes",              "Yes"),
    ("Bank transfer accepted",       "No",          "Optional",         "Yes",              "Yes"),
    ("Custom subdomain",             "No",          "No",               "Optional",         "Yes"),
    ("White-label report exports",   "No",          "Yes",              "Yes",              "Yes"),
    ("Dedicated account manager",    "No",          "No",               "Yes",              "Yes"),
    ("Slack/Teams integration",      "No",          "No",               "Yes",              "Yes"),
    ("Custom enrichment modes",      "No",          "No",               "No",               "Yes"),
    ("On-premise / private deploy",  "No",          "No",               "No",               "Optional"),
    ("SLA (99.9% uptime)",           "No",          "No",               "Yes",              "Yes"),
    ("Security review / NDA",        "No",          "No",               "Optional",         "Yes"),
]

plan_colors_corp = [Z800, Z900, RED, Z900, BLACK]
for ri, row_data in enumerate(corp_features, 2):
    for ci, val in enumerate(row_data, 1):
        bg = plan_colors_corp[ci-1] if ri == 2 else (Z900 if ri%2==0 else BLACK)
        bold = ri == 2
        color = WHITE if ri == 2 else (GREEN if val == "Yes" else (RED+"88" if val == "No" else (AMBER if val not in ["—","No","Yes"] else Z400)))
        cell_style(ws4, ri, ci, val, bold=bold, size=10 if ri>2 else 11, color=color, bg=bg, align="center" if ci>1 else "left")
    ws4.row_dimensions[ri].height = 22

# Pricing for corp
ws4.row_dimensions[len(corp_features)+3].height = 32
pr = len(corp_features) + 3
ws4.merge_cells(f"A{pr}:E{pr}")
cell_style(ws4, pr, 1, "── CORPORATE PRICING (Annual Contracts — 2 months free) ──", bold=True, size=11, color=RED, bg=Z900)

corp_pricing = [
    ("Agency (5 seats)",      "KES 15,000/mo",  "KES 150,000/yr",  "~$1,154/yr",    "Per extra seat: KES 2,500/mo"),
    ("Enterprise (10 seats)", "KES 30,000/mo",  "KES 300,000/yr",  "~$2,308/yr",    "Per extra seat: KES 2,000/mo"),
    ("Enterprise (20 seats)", "KES 55,000/mo",  "KES 550,000/yr",  "~$4,231/yr",    "Per extra seat: KES 1,500/mo"),
    ("Enterprise+",           "Custom",          "Custom",           "Custom",        "Contact: dr.dullu@outlook.com"),
]

for ri2, row_data in enumerate(corp_pricing, pr+2):
    for ci, val in enumerate(row_data, 1):
        bg = Z900 if ri2%2==0 else BLACK
        color = AMBER if "KES" in val or "$" in val else (GREEN if "Contact" in val else (WHITE if ci==1 else Z200))
        cell_style(ws4, ri2, ci, val, size=10, color=color, bg=bg, align="left")
    ws4.row_dimensions[ri2].height = 22

# ── Final save ────────────────────────────────────────────────────────────────
output = "/home/dullz/hunter-saas/hunter-ad/hunter_pricing_plans.xlsx"
wb.save(output)
print(f"Saved: {output}")
